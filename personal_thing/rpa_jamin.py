import pyautogui
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from imap_tools import MailBox
from account import *
from openpyxl import Workbook
import smtplib
from email.message import EmailMessage
from typing import Text
from speech_recognition import *
import time
from pyautogui import *
import clipboard
import keyboard
import pyaudio
from gtts import gTTS
import playsound


# 마이크 함수 설정
def jamin_mic(text1):
    recognizer = Recognizer()
    mic = Microphone()

    with mic as source:
        speak(text1)

        audio = recognizer.listen(source)

        try:
            text = recognizer.recognize_google(audio, language='ko')
            print(text)

        except:
            speak("인식하지 못했어요")

    return text
    


# 음성 출력
def speak(text1):
    tts = gTTS(text=text1, lang='ko')
    filename = 'voice.wav'
    tts.save(filename)
    playsound.playsound(filename)




mail_list = []



# 하고자 하는 작업 입력받기
# result1 = pyautogui.prompt("하고자 하는 작업번호을 입력하시오\n1. naver\n2. google\n3. youtube\n4. excel\n5. email", "입력")


result1 = jamin_mic("원하시는 작업을 말씀해주세요.")


if result1 == '네이버':
    result2 = jamin_mic("하고자 하는 작업을 말씀해주세요 스포츠 증권 검색")
    if result2 == '스포츠':
        result2 = "스포츠"
        browser = webdriver.Chrome('./personal_thing/chromedriver')
        browser.get('http://naver.com')
        elem = browser.find_element_by_link_text(result2)
        elem.click()

    elif result2 == '증권':
        result2 = "증권"
        browser = webdriver.Chrome('./personal_thing/chromedriver')
        browser.get('http://naver.com')
        elem = browser.find_element_by_link_text(result2)
        elem.click()
        
    elif result2 == '검색':
        result = jamin_mic("검색할 내용을 말씀해주세요")
        browser = webdriver.Chrome('./personal_thing/chromedriver')
        browser.get('http://naver.com')
        elem = browser.find_element_by_xpath('//*[@id="query"]')
        elem.send_keys(result)
        elem.send_keys(Keys.ENTER)
    else:
        pyautogui.alert("해당 작업 없음", "결과")



elif result1 == '구글':
    result3 = jamin_mic("검색하고자하는 내용을 말씀해주세요")
    browser = webdriver.Chrome('./personal_thing/chromedriver')
    browser.get('http://google.com')
    elem = browser.find_element_by_xpath('/html/body/div[1]/div[3]/form/div[1]/div[1]/div[1]/div/div[2]/input')
    elem.send_keys(result3)
    elem.send_keys(Keys.ENTER)

elif result1 == '유튜브':
    result4 = jamin_mic("검색하고자하는 내용을 말씀해주세요")
    browser = webdriver.Chrome('./personal_thing/chromedriver')
    browser.get('http://youtube.com')
    elem = browser.find_element_by_xpath('//*[@id="search"]')
    elem.send_keys(result4)
    time.sleep(1)
    elem.send_keys(Keys.ENTER)

elif result1 == '엑셀':
    result5 = pyautogui.prompt("작업하고자 하는 엑셀을 입력하세요", "입력")
    wb = load_workbook(result5 + ".xlsx")
    ws = wb.active
    result6 = pyautogui.prompt("1. 수정\n2. 출력", "입력")
    if result6 == '1':
        row1 = pyautogui.prompt("수정 품명을 입력하세요", "입력")
        column1 = pyautogui.prompt("수정 항목을 입력하세요", "입력")
        change_one = pyautogui.prompt("수정 내용을 입력하세요", "입력")

        xy1 = None
        xy2 = None

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", horizontal="center")
                if cell.value == row1:
                    xy1 = coordinate_from_string(cell.coordinate)
                    print("품명행"+ str(xy1[1]))
                elif cell.value == column1:
                    xy2 = coordinate_from_string(cell.coordinate)
                    print("항목열"+ str(xy2[0]))



        ws[str(xy2[0]) + str(xy1[1])] = change_one

    if result6 == '2':
        res1 = pyautogui.prompt("출력 품명을 입력하세요", "입력")
        res2 = pyautogui.prompt("출력 항목을 입력하세요", "입력")
        
        xy1 = None
        xy2 = None

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", horizontal="center")
                if cell.value == res1:
                    xy3 = coordinate_from_string(cell.coordinate)
                elif cell.value == res2:
                    xy4 = coordinate_from_string(cell.coordinate)



        pyautogui.alert("품명"+res1+"의 "+res2+"은 "+str(ws[str(xy4[0])+str(xy3[1])].value)+ " 입니다", "결과")
        
                

    wb.save(result5 + ".xlsx")


elif result1 == '이메일':
    re = pyautogui.prompt("검색할 내용을 입력하시오", "입력")

    with MailBox("imap.naver.com", 993).login("bamer@naver.com", "mok1670!", initial_folder="INBOX") as mailbox:
        for msg in mailbox.fetch(reverse=True):
            if re in msg.from_:
                print("송신: {}, 제목: {}".format(msg.from_, msg.subject))
                mail_list.append((msg.from_, msg.subject, msg.text))
                




else:
    pyautogui.alert("해당 작업 없음", "결과")







