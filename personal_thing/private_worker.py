import pyautogui
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from imap_tools import MailBox
from account import *
from openpyxl import Workbook
import smtplib
from email.message import EmailMessage
from typing import Text
from speech_recognition import *
import time
from pyautogui import *
import clipboard
import keyboard
import pyaudio
from gtts import gTTS
import playsound

# 마이크 함수 설정
def jamin_mic(text1):
    recognizer = Recognizer()
    mic = Microphone()

    with mic as source:
        speak(text1)

        audio = recognizer.listen(source)

        try:
            text = recognizer.recognize_google(audio, language='ko')
            print(text)

        except:
            speak("인식하지 못했어요")

    return text

# 음성 출력
def speak(text1):
    tts = gTTS(text=text1, lang='ko')
    filename = 'voice.wav'
    tts.save(filename)
    playsound.playsound(filename)


result1 = jamin_mic("재민님 무엇을 도와드릴까요? 검색, 메일, 엑셀, 쇼핑, 스크래핑, 일정 작업이 있습니다.")


if result1 == "일정":
    result1_1 = jamin_mic("날짜를 말씀해주세요.")
    wb = load_workbook("일정.xlsx")
    ws = wb.active
    for row in ws.iter_rows(max_row=1):
        for cell in row:
            if cell.value == result1_1:
                xy4 = coordinate_from_string(cell.coordinate)

    cells = ws[str(xy4[0])]
    for cell in cells:
        speak(cell.value)

elif result1 == "검색":
    result2_1 = jamin_mic("검색 사이트를 말씀해주세요. 유튜브, 네이버, 구글이 있습니다.")   

elif result1 == "메일":
    print("a")

elif result1 == "엑셀":
    print("a")

elif result1 == "쇼핑":
    result5_1 = jamin_mic("찾고자하는 품목을 말씀해주세요")

else:
    speak("해당 작업이 없습니다.")

