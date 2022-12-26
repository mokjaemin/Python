from imap_tools import MailBox
from account import *
from openpyxl import Workbook
import smtplib
from email.message import EmailMessage




applicant_list = [] # 지원 명단
success_list = [] # 당첨 명단
loser_list = [] # 탈락 명단
success = 3 # 선정자 수

# 메일 정보 불러오기
with MailBox("imap.gmail.com", 993).login(EMAIL_ADDRESS, EMAIL_PASSWORD, initial_folder="INBOX") as mailbox:
    idx = 1
    for msg in mailbox.fetch():
        if "파이썬 특강 신청합니다." in msg.subject:
            name, phone = msg.text.split("/")
            # print("순번: {}, 이름: {}, 이메일: {}, 번호: {}".format(idx, name, msg.from_, phone))
            applicant_list.append((idx, name, msg.from_, phone))
            idx += 1

# 합격자 명단 작성
for i in range(0, success):
    # print(applicant_list[i])
    success_list.append(applicant_list[i])


# 틸락자 명단 작성
for i in range(success, len(applicant_list)):
    # print(applicant_list[i])
    loser_list.append(applicant_list[i])


# 합격자 엑셀 생성
wb = Workbook() #새 워크북 생성 
ws = wb.active #현재 활성화된 시트를 가져옴

for x in range(1, success+1):
    ws.cell(row = x, column = 1, value = success_list[x-1][1])
for x in range(1, success+1):
    ws.cell(row = x, column = 2, value = success_list[x-1][3])

wb.save("james.xlsx")



# 합격자 메일 전송
for i in range(0, success):
    with smtplib.SMTP('smtp.gmail.com', 587) as smtp:

        email1 = success_list[i][2]
        smtp.ehlo() # 연결이 잘 수립되었는 지 확인
        smtp.starttls() # 모든 내용이 암호화되어 전송
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD) # 로그인

        subject = "합격 메일" # 메일 제목
        body = "합격하셨습니다 " + "순번 : " + str(success_list[i][0])# 메일 본문

        msg = EmailMessage()
        msg["Subject"] = subject
        msg["To"] = success_list[i][2]
        msg["From"] = EMAIL_ADDRESS
        msg.set_content(body)
        smtp.send_message(msg)


# 탈락자 메일 전송
for i in range(0, len(applicant_list)-success):
    with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
        
        smtp.ehlo() # 연결이 잘 수립되었는 지 확인
        smtp.starttls() # 모든 내용이 암호화되어 전송
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD) # 로그인

        subject = "탈락 메일" # 메일 제목
        body = "탈락하셨습니다 " + "순번 : " + str(loser_list[i][0]-3)# 메일 본문

        msg = EmailMessage()
        msg["Subject"] = subject
        msg["To"] = loser_list[i][2]
        msg["From"] = EMAIL_ADDRESS
        msg.set_content(body)
        smtp.send_message(msg)
