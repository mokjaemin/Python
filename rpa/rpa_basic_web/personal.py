import pyautogui
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from imap_tools import MailBox
from account import *
from openpyxl import Workbook
import smtplib
from email.message import EmailMessage



mail_list = []



# 하고자 하는 작업 입력받기
result1 = pyautogui.prompt("하고자 하는 작업번호을 입력하시오\n1. naver\n2. google\n3. youtube\n4. excel\n5. email", "입력")



if result1 == '1':
    result2 = pyautogui.prompt("하고자 하는 작업번호을 입력하시오\n1. 스포츠\n2. 증권\n3. 검색\n", "입력")
    if result2 == '1':
        result2 = "스포츠"
        browser = webdriver.Chrome('./chromedriver')
        browser.get('http://naver.com')
        elem = browser.find_element_by_link_text(result2)
        elem.click()

    elif result2 == '2':
        result2 = "증권"
        browser = webdriver.Chrome('./chromedriver')
        browser.get('http://naver.com')
        elem = browser.find_element_by_link_text(result2)
        elem.click()
        
    elif result2 == '3':
        result = pyautogui.prompt("검색할 내용을 입력하시오", "입력")
        browser = webdriver.Chrome('./chromedriver')
        browser.get('http://naver.com')
        elem = browser.find_element_by_xpath('//*[@id="query"]')
        elem.send_keys(result)
        elem.send_keys(Keys.ENTER)
    else:
        pyautogui.alert("해당 작업 없음", "결과")



elif result1 == '2':
    result3 = pyautogui.prompt("검색하고자하는 내용을 입력하시오", "입력")
    browser = webdriver.Chrome('./chromedriver')
    browser.get('http://google.com')
    elem = browser.find_element_by_xpath('/html/body/div[1]/div[3]/form/div[1]/div[1]/div[1]/div/div[2]/input')
    elem.send_keys(result3)
    elem.send_keys(Keys.ENTER)

elif result1 == '3':
    result4 = pyautogui.prompt("검색하고자하는 내용을 입력하시오", "입력")
    browser = webdriver.Chrome('./chromedriver')
    browser.get('http://youtube.com')
    elem = browser.find_element_by_xpath('//*[@id="search"]')
    elem.send_keys(result4)
    time.sleep(1)
    elem.send_keys(Keys.ENTER)

elif result1 == '4':
    result5 = pyautogui.prompt("작업하고자 하는 엑셀을 입력하세요", "입력")
    wb = load_workbook(result5 + ".xlsx")
    ws = wb.active
    result6 = pyautogui.prompt("1. 수정\n2. 출력", "입력")
    if result6 == '1':
        row1 = pyautogui.prompt("수정 품명을 입력하세요", "입력")
        column1 = pyautogui.prompt("수정 항목을 입력하세요", "입력")
        change_one = pyautogui.prompt("수정 내용을 입력하세요", "입력")

        xy1 = None
        xy2 = None

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", horizontal="center")
                if cell.value == row1:
                    xy1 = coordinate_from_string(cell.coordinate)
                    print("품명행"+ str(xy1[1]))
                elif cell.value == column1:
                    xy2 = coordinate_from_string(cell.coordinate)
                    print("항목열"+ str(xy2[0]))



        ws[str(xy2[0]) + str(xy1[1])] = change_one

    if result6 == '2':
        res1 = pyautogui.prompt("출력 품명을 입력하세요", "입력")
        res2 = pyautogui.prompt("출력 항목을 입력하세요", "입력")
        
        xy1 = None
        xy2 = None

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", horizontal="center")
                if cell.value == res1:
                    xy3 = coordinate_from_string(cell.coordinate)
                elif cell.value == res2:
                    xy4 = coordinate_from_string(cell.coordinate)



        pyautogui.alert("품명"+res1+"의 "+res2+"은 "+str(ws[str(xy4[0])+str(xy3[1])].value)+ " 입니다", "결과")
        
                

    wb.save(result5 + ".xlsx")


elif result1 == '5':
    re = pyautogui.prompt("검색할 내용을 입력하시오(제목, 메일, 내용 상관없습니다)\n최신 메일을 보여줍니다.", "입력")

    with MailBox("imap.naver.com", 993).login("bamer@naver.com", "mok1670!", initial_folder="INBOX") as mailbox:
        for msg in mailbox.fetch(reverse=True):
            if re in msg.from_:
                print("송신: {}, 제목: {}".format(msg.from_, msg.subject))
                mail_list.append((msg.from_, msg.subject, msg.text))
                




else:
    pyautogui.alert("해당 작업 없음", "결과")








