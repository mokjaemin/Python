from openpyxl import load_workbook #파일 불러오기
wb = load_workbook("james.xlsx") # 해당엑셀 불러옴
ws = wb.active # 현재 활성화된 시트를 가져옴

# 셀 데이터 불러오기
for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(row=x, column=y).value, end= " ")
    print() # 줄바꿈


# 셀 갯수를 모를 때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end= " ")
    print()
