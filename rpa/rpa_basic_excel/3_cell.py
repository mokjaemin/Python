from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "James"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

print(ws["A1"].value) #A1셀의 값을 출력, 값이 없을땐 none을 출력

#row = 1, 2, 3 ...
#column = A(1), B(2), C(3) ...
print(ws.cell(row = 1, column=1).value) #A1의 값
ws.cell(row = 1, column=2, value = 10) #해당 셀의 값을 집어넣음




#반복문을 이용해서 random 숫자 대입
index = 0
from random import *
for x in range(1, 11): #10개의 row
    for y in range(1, 11): #10개의 column
        #ws.cell(row = x, column = y, value = randint(0,100)) #랜덤한 숫자 채우기
        ws.cell(row=x, column=y, value=index)
        index += 1







wb.save("james.xlsx")