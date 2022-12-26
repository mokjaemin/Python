from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws.merge_cells("B2:D2") #비투부터 디투까지 합병
ws["B2"].value = "Merged Cell"

wb.save("james_merge.xlsx")