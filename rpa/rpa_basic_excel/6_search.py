from openpyxl import load_workbook
from openpyxl.reader.excel import load_workbook
wb = load_workbook("example.xlsx")
ws = wb.active


# min_row = 2 는 첫번째 행 제외, row[1]은 두번째 열 해당
for row in ws.iter_rows(min_row=2):
    if int(row[1].value)>70:
        print(row[0].value, "번학생은 고득점자")

for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "번호":
            cell.value = "product"
        if cell.value == "영어":
            cell.value = "input"
        if cell.value == "수학":
            cell.value = "output"

ws["A2"] = "A"
ws["A3"] = "B"
ws["A4"] = "C"
ws["A5"] = "D"
ws["A6"] = "E"
ws["A7"] = "F"
ws["A8"] = "G"
ws["A9"] = "H"
ws["A10"] = "I"
ws["A11"] = "J"



wb.save("example.xlsx")