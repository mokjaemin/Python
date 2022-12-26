import random
from openpyxl import load_workbook
wb = load_workbook("james.xlsx")
ws = wb.active

# ws.insert_cols(2)
# ws["B1"] = "국어"
# for i in range(2, 12):
#     ws.cell(row = i, column=2, value = random.randint(0, 100)) #해당 셀의 값을 집어넣음
# # 내가 쓴거    

ws.move_range("B1:C11", rows=0, cols=1) # 비원 기준으로 줄 = 줄 + 로우스, 컬럼 = 컬럼 + 컬럼
ws["B1"].value = "국어"



wb.save("james_move.xlsx")