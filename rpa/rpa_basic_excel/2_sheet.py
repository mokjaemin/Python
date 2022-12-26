from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() #새로운 시트 기본 이름으로 생성
ws.title = "james2" #시트 이름 변경
ws.sheet_properties.tabColor = "ff66ff" #시트 색깔 변경

ws1 = wb.create_sheet("james3") #이름과 함께 시트 생성
ws2 = wb.create_sheet("james4", 3) #해당 인덱스에 시트 생성

print(wb.sheetnames) #시트이름 확인

#시트 복사
ws["A1"] = "test"
target = wb.copy_worksheet(ws)
target.title = "copied sheet"

wb.save("james.xlsx")
