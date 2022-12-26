
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook("james.xlsx")
ws = wb.active


a1 = ws["A1"] #번호
b1 = ws["B1"] #영어
c1 = ws["C1"] #수학

#A의 열의 너비 설정
ws.column_dimensions["A"].width = 5
# 높이 - ws.row_dimensions[1].height


# 스타일 적용
a1.font = Font(color="FF0000", italic=True, bold=True, name="Arial", strike=True) 
#글자색은 빨갛게, 이탤릭(살짝 눕히기), 두껍게, 글자체, 취소선, 크기=size, 밑줄=> underline="single" or "double"


# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점 이상의 학생들을 초록색으로 표시
for row in ws.rows:
    for cell in row:
        # 각 셀에 대한 중앙 정렬
        cell.alignment = Alignment(vertical="center", horizontal="center")
        
        if cell.column == 1:
            continue # A번호열 제외
        # 셀의 값이 int형이면서 크기가 90이상
        if isinstance(cell.value, int) and cell.value > 80:
            cell.fill = PatternFill(fgColor= "00FF00", fill_type= "solid")
            # 배경을 초록색으로 설정
            cell.font = Font(color="FF0000")

# 틀 고정(스크롤 내려도 지정행은 고정됨)
ws.freeze_panes = "B2" # 비투 기준으로 스크롤내려도 고정


wb.save("james_style.xlsx")