from openpyxl import Workbook
from random import *
from openpyxl.utils import cell


wb = Workbook()
ws = wb.active

# 한줄 씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11): # 10개의 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])


col_B = ws["B"] # 영어 점수 가져옴
#for cell in col_B:
#    print(cell.value)

col_range = ws["B:C"] # B부터 C까지 가져옴 영어부터 수학
#for cell1 in col_range:
#    for cell2 in cell1:
#        print(cell2.value)


row_title = ws[1] # 첫번째 row만 출력
#for cell in row_title:
#   print(cell.value)

row_range = ws[2:6] # 두번째 줄에서 6번째 줄까지 출력
#for cell in row_range:
#    for cell1 in cell:
#        print(cell1.value, end=" ")
#    print()

# 셀의 좌표정보를 가져오는 임포트
from openpyxl.utils.cell import coordinate_from_string, coordinate_to_tuple

row_range1 = ws[2:ws.max_row] # 두번째부터 끝까지
for rows in row_range1:
    for rows1 in rows:
        #print(rows1.value, end=" ")
        #print(rows1.coordinate, end = " ") # 셀의 좌표정보 가져오기
        xy = coordinate_from_string(rows1.coordinate) # 셀의 정보를 끊어서 보여줌
        print(xy, end=" ") #xy[0] - A, B... xy[1] - 1,2,3...

    print()


# 전체 rows
# # print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[1].value) #특정 줄 가져오기 컬럼은 columns로 바꿔주면 됨
# for row in ws.iter_rows(): #전체를 불러옴 컬럼은 cols
#     print(row[1].value)

# 특정 범위 출력
# for row in ws.iter_rows(min_row=1, max_row=11, min_col=2, max_col=3):
#     print(row[0].value, row[1].value)



# 전체 columns
# print(tuple(ws.columns))






wb.save("james.xlsx")
