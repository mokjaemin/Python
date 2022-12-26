from openpyxl import Workbook
import datetime
wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today()
ws["A2"] = "=SUM(1,2,3)" # 1,2,3 이 더해진 값
ws["A3"] = "=AVERAGE(1,2,3)" # 1,2,3 의 평균

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)" 

wb.save("formula.xlsx")