from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active


ws.append(["번호", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"])
scores = [
(1,10,8,5,14,26,12),
(2,7,3,7,15,24,18),
(3,9,5,8,8,12,4),
(4,7,8,7,17,21,18),
(5,7,8,7,16,25,15),
(6,3,5,8,8,17,0),
(7,4,9,10,16,27,18),
(8,6,6,6,15,19,17),
(9,10,10,9,19,30,19),
(10,9,8,8,20,25,20)
]

for s in scores:
    ws.append(s)


# col_D = ws["D2:D11"]
# for cell1 in col_D:
#     for cell2 in cell1:
#         cell2.value = 10

# or

for idx, cell in enumerate(ws["D"]):
    if idx == 0:
        continue
    cell.value = 10

# ws["I1"] = "총점"
# ws["I2"] = "=SUM(B2:G2)"
# ws["I3"] = "=SUM(B3:G3)"
# ws["I4"] = "=SUM(B4:G4)"
# ws["I5"] = "=SUM(B5:G5)"
# ws["I6"] = "=SUM(B6:G6)"
# ws["I7"] = "=SUM(B7:G7)"
# ws["I8"] = "=SUM(B8:G8)"
# ws["I9"] = "=SUM(B9:G9)"
# ws["I10"] = "=SUM(B10:G10)"
# ws["I11"] = "=SUM(B11:G11)"

for idx, score in enumerate(scores, start=2):
    sum_val = sum(score[1:],) - score[3] + 10
    ws.cell(row = idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx)


    grade = None
    if sum_val >= 90:
        grade = "A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"
    
    
    if score[1] < 5:
        grade = "F"

    ws.cell(row = idx, column=9).value = grade

ws["H1"] = "총점"
ws["I1"] = "성적"
# for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=8):
#     result = row[0].value + row[1].value + row[2].value + row[3].value + row[4].value + row[5].value
#     print(result)
#     if result >= 60:
#         row[6].value = "A"
#     if result >= 50 and result < 60:
#         row[6].value = "B"
#     if result >= 40 and result < 50:
#         row[6].value = "C"
#     if result < 40:
#         row[6].value = "D"
#     if row[0].value < 5:
#         row[6].value = "F"



wb.save("성적표.xlsx")